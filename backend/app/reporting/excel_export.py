"""Exportación del informe a Excel (.xlsx).

Dos hojas: 'Resumen' (cabecera + una fila por cuenta) y 'Detalle' (un apunte por
fila, con su clasificación de cuenta). Pensado para que un humano audite.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ..domain.models import SUBCATEGORIA_INFO
from ..domain.resultados import Informe

_COLOR = {
    "SIN_FACTURA_ALTA_CONFIANZA": "FFC7CE",  # rojo claro
    "REVISAR": "FFEB9C",                      # ámbar
    "CONCILIADA": "C6EFCE",                   # verde
    "NO_FIABLE": "D9D9D9",                    # gris
    "EXCLUIDA": "F2F2F2",                     # gris muy claro
}


def exportar_excel(inf: Informe) -> bytes:
    wb = Workbook()

    # --- Hoja Resumen ------------------------------------------------------
    ws = wb.active
    ws.title = "Resumen"
    neg = Font(bold=True)
    ws.append(["Informe de pagos sin factura — v1 (precision-first)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    res = inf.resumen
    ws.append(["Pagos SIN factura (alta confianza)", res.n_sin_factura,
               f"{res.importe_sin_factura} €"])
    ws.append(["En revisión (REVISAR)", res.n_revisar, f"{res.importe_revisar} €"])
    ws.append(["Conciliadas", res.n_conciliadas])
    ws.append(["No fiables", res.n_no_fiables])
    ws.append(["Excluidas (técnicas)", res.n_excluidas])
    ws.append(["Total cuentas", res.n_cuentas])
    if inf.flags_globales:
        ws.append([])
        ws.append(["Avisos:", ", ".join(inf.flags_globales)])
    ws.append([])

    cab = ["Código", "Cuenta", "Clasificación", "Subcasilla (REVISAR)", "Confianza",
           "Σ Debe", "Σ Haber", "Saldo recon.", "€ sospechoso", "Nº fact.",
           "Nº pagos", "Nº abonos", "Flags", "Motivo"]
    fila_cab = ws.max_row + 1
    ws.append(cab)
    for c in ws[fila_cab]:
        c.font = neg
    for r in inf.resultados:
        sub = SUBCATEGORIA_INFO.get(r.subcategoria, {}).get("etiqueta", "") if r.subcategoria else ""
        motivo = r.motivo + ((" || " + r.subcategoria_motivo) if r.subcategoria_motivo else "")
        ws.append([
            r.codigo_cuenta, r.nombre_cuenta, r.clasificacion.value, sub,
            r.confianza.value, float(r.suma_debe), float(r.suma_haber),
            float(r.saldo_reconstruido), float(r.importe_sospechoso),
            r.num_facturas, r.num_pagos, r.num_abonos,
            ", ".join(r.flags), motivo,
        ])
        fill = _COLOR.get(r.clasificacion.value)
        if fill:
            ws.cell(row=ws.max_row, column=3).fill = PatternFill(
                "solid", fgColor=fill)
    _ajustar(ws)

    # --- Hoja Detalle ------------------------------------------------------
    wd = wb.create_sheet("Detalle")
    cab2 = ["Código", "Cuenta", "Clasif. cuenta", "Fecha", "Asiento", "Tipo",
            "Comentario", "Debe", "Haber", "Referencia", "Saldo"]
    wd.append(cab2)
    for c in wd[1]:
        c.font = neg
    for r in inf.resultados:
        for m in r.movimientos:
            ref = m.referencias.su_factura or m.referencias.factura or ""
            wd.append([
                r.codigo_cuenta, r.nombre_cuenta, r.clasificacion.value,
                m.fecha.isoformat() if m.fecha else "", m.asiento, m.tipo.value,
                m.comentario, float(m.debe), float(m.haber), ref,
                float(m.saldo_reportado) if m.saldo_reportado is not None else "",
            ])
    _ajustar(wd)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _ajustar(ws) -> None:
    for col in ws.columns:
        ancho = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        letra = col[0].column_letter
        ws.column_dimensions[letra].width = min(ancho + 2, 60)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].alignment = Alignment(wrap_text=True)
